"""
Excel Parser Service

@SPEC:DASH-001
@CODE:EXCEL-PARSER

Parses Excel files (.xlsx, .xls) and extracts data into structured records.
"""

from typing import Dict, List, Any, BinaryIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ExcelParserService:
    """
    Service for parsing Excel files into structured data.

    @CODE:EXCEL-PARSER-SERVICE

    Methods:
        parse(file, has_header): Parse Excel file and return records
        get_metadata(file): Get file metadata (sheet count, dimensions)
    """

    def parse(self, file: BinaryIO, has_header: bool = True) -> Dict[str, Any]:
        """
        Parse an Excel file and extract records.

        @CODE:EXCEL-PARSER-001

        Args:
            file: Binary file object (BytesIO or file handle)
            has_header: If True, treat first row as headers

        Returns:
            Dict with structure:
            - success (bool): Whether parsing succeeded
            - records (List[Dict]): List of data records
            - headers (List[str]): Column headers
            - error (str, optional): Error message if failed

        Business Rules:
        - Parse first sheet only
        - Convert all data types to Python native types
        - Handle empty cells as None
        - Validate file is not empty
        """
        try:
            # Load workbook
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active

            if ws is None:
                return {'success': False, 'error': 'Workbook has no active sheet'}

            # Get all rows as list
            rows = list(ws.rows)

            if not rows:
                return {'success': False, 'error': 'Excel file is empty'}

            headers = []
            data_rows = []

            if has_header:
                # First row is headers
                header_row = rows[0]
                headers = [self._cell_to_value(cell) for cell in header_row]

                # Remaining rows are data
                data_rows = rows[1:]

            else:
                # No headers - use column indices
                if rows:
                    num_cols = len(rows[0])
                    headers = [f'col_{i}' for i in range(num_cols)]
                    data_rows = rows

            # Convert rows to dictionaries
            records = []
            for row in data_rows:
                # Skip completely empty rows
                values = [self._cell_to_value(cell) for cell in row]
                if all(v is None for v in values):
                    continue

                # Create record dict
                record = {}
                for header, value in zip(headers, values):
                    record[str(header)] = value

                records.append(record)

            wb.close()

            return {
                'success': True,
                'records': records,
                'headers': headers
            }

        except InvalidFileException:
            return {'success': False, 'error': 'Invalid Excel file format'}

        except Exception as e:
            return {'success': False, 'error': f'Failed to parse Excel file: {str(e)}'}

    def get_metadata(self, file: BinaryIO) -> Dict[str, Any]:
        """
        Get metadata about an Excel file.

        @CODE:EXCEL-PARSER-002

        Args:
            file: Binary file object

        Returns:
            Dict with metadata:
            - sheet_count: Number of sheets
            - row_count: Number of rows in active sheet
            - column_count: Number of columns in active sheet
            - sheet_names: List of sheet names
        """
        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active

            # Get dimensions
            row_count = ws.max_row
            column_count = ws.max_column

            metadata = {
                'sheet_count': len(wb.sheetnames),
                'row_count': row_count,
                'column_count': column_count,
                'sheet_names': wb.sheetnames
            }

            wb.close()

            return metadata

        except Exception as e:
            return {
                'sheet_count': 0,
                'row_count': 0,
                'column_count': 0,
                'error': str(e)
            }

    def _cell_to_value(self, cell) -> Any:
        """
        Convert an Excel cell to a Python value.

        @CODE:EXCEL-PARSER-003

        Args:
            cell: openpyxl Cell object

        Returns:
            Python native type (str, int, float, datetime, None)
        """
        value = cell.value

        if value is None:
            return None

        # Handle datetime
        if isinstance(value, datetime):
            return value.isoformat()

        # Return as-is for other types
        return value
