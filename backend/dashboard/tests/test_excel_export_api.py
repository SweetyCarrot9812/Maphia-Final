"""
Tests for Excel Export API functionality.

@SPEC:EXPORT-001 - Excel Export System

Test Coverage:
- REQ-EXPORT-003: Basic Excel export with styling
- REQ-EXPORT-004: Multi-sheet support (Data, Statistics, Charts)
- Header styling (bold, colored background)
- Auto-width columns
- Zebra striping (alternating row colors)
- Role-based access control
"""

import pytest
import openpyxl
from io import BytesIO
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APIClient
from dashboard.models import Dataset, DataRecord
from django.contrib.auth import get_user_model

User = get_user_model()


@pytest.mark.django_db
class TestExcelExportAPI:
    """Test suite for Excel export functionality."""

    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test data before each test."""
        self.client = APIClient()

        # Create test users
        self.admin_user = User.objects.create_user(
            username='admin',
            email='admin@test.com',
            password='testpass123',
            role='admin'
        )
        self.manager_user = User.objects.create_user(
            username='manager',
            email='manager@test.com',
            password='testpass123',
            role='manager'
        )
        self.viewer_user = User.objects.create_user(
            username='viewer',
            email='viewer@test.com',
            password='testpass123',
            role='viewer'
        )

        # Create test dataset
        self.dataset = Dataset.objects.create(
            title='Test Excel Dataset',
            description='Test dataset for Excel export',
            filename='test.xlsx',
            file_size=2048,
            record_count=5,
            category='enrollment',
            uploaded_by=self.admin_user
        )

        # Create test records with various data types
        self.records = [
            DataRecord.objects.create(
                dataset=self.dataset,
                data={
                    'student_name': '홍길동',
                    'age': 25,
                    'gpa': 3.85,
                    'enrolled': True,
                    'category': 'enrollment'
                }
            ),
            DataRecord.objects.create(
                dataset=self.dataset,
                data={
                    'student_name': 'Jane Smith',
                    'age': 30,
                    'gpa': 3.92,
                    'enrolled': True,
                    'category': 'enrollment'
                }
            ),
            DataRecord.objects.create(
                dataset=self.dataset,
                data={
                    'student_name': 'Bob Johnson',
                    'age': 22,
                    'gpa': 3.45,
                    'enrolled': False,
                    'category': 'enrollment'
                }
            ),
            DataRecord.objects.create(
                dataset=self.dataset,
                data={
                    'student_name': 'Alice Lee',
                    'age': 28,
                    'gpa': 3.78,
                    'enrolled': True,
                    'category': 'enrollment'
                }
            ),
            DataRecord.objects.create(
                dataset=self.dataset,
                data={
                    'student_name': 'Charlie Brown',
                    'age': 24,
                    'gpa': 3.60,
                    'enrolled': True,
                    'category': 'enrollment'
                }
            ),
        ]

    def test_excel_export_requires_authentication(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that Excel export requires authentication.

        Expected: 401 Unauthorized
        """
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_excel_export_admin_access(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that admin users can export Excel.

        Expected: 200 OK with Excel file
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'attachment; filename=' in response['Content-Disposition']

    def test_excel_export_manager_access(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that manager users can export Excel.

        Expected: 200 OK with Excel file
        """
        self.client.force_authenticate(user=self.manager_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

    def test_excel_export_viewer_denied(self):
        """
        Test that viewer users cannot export Excel.

        Expected: 403 Forbidden
        """
        self.client.force_authenticate(user=self.viewer_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_excel_export_valid_workbook(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that exported file is a valid Excel workbook.

        Expected: File can be opened with openpyxl
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        # Try to load the Excel file
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        assert workbook is not None

    def test_excel_export_multiple_sheets(self):
        """
        @SPEC:REQ-EXPORT-004
        Test that Excel export contains multiple sheets.

        Expected:
        - Sheet 1: "데이터" (Data)
        - Sheet 2: "요약 통계" (Summary Statistics)
        - Sheet 3: "차트" (Charts) - optional based on data
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet_names = workbook.sheetnames

        assert '데이터' in sheet_names or 'Data' in sheet_names
        assert '요약 통계' in sheet_names or 'Summary' in sheet_names or 'Statistics' in sheet_names

    def test_excel_export_data_sheet_content(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that data sheet contains all records.

        Expected: All 5 records present
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        data_sheet = workbook.active

        # Check row count (header + 5 data rows)
        assert data_sheet.max_row >= 6

        # Check header row exists
        first_row = [cell.value for cell in data_sheet[1]]
        assert 'student_name' in first_row or any('name' in str(val).lower() for val in first_row if val)

    def test_excel_export_header_styling(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that header row has proper styling.

        Expected:
        - Bold font
        - Colored background (#4472C4 or similar blue)
        - White text color
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        data_sheet = workbook.active

        # Check first cell styling (header)
        first_cell = data_sheet['A1']

        # Check bold font
        assert first_cell.font.bold is True

        # Check background fill color (should have a fill pattern)
        assert first_cell.fill.patternType is not None
        assert first_cell.fill.patternType != 'none'

    def test_excel_export_auto_width_columns(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that columns have auto-adjusted width.

        Expected: Column widths are set (not default 8.43)
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        data_sheet = workbook.active

        # Check at least one column has adjusted width
        adjusted_columns = [
            col for col in data_sheet.column_dimensions.values()
            if col.width != 8.43  # Default width
        ]

        assert len(adjusted_columns) > 0

    def test_excel_export_zebra_striping(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that data rows have zebra striping (alternating colors).

        Expected: Odd and even rows have different background colors
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        data_sheet = workbook.active

        # Check second and third data rows (after header)
        if data_sheet.max_row >= 3:
            row2_fill = data_sheet['A2'].fill
            row3_fill = data_sheet['A3'].fill

            # At least one should have a fill (zebra striping applied)
            has_striping = (
                row2_fill.patternType != 'none' or
                row3_fill.patternType != 'none'
            )

            assert has_striping

    def test_excel_export_statistics_sheet(self):
        """
        @SPEC:REQ-EXPORT-004
        Test that statistics sheet contains aggregated data.

        Expected: Statistics sheet with summary information
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))

        # Find statistics sheet
        stats_sheet = None
        for sheet_name in workbook.sheetnames:
            if '통계' in sheet_name or 'statistic' in sheet_name.lower() or 'summary' in sheet_name.lower():
                stats_sheet = workbook[sheet_name]
                break

        assert stats_sheet is not None
        assert stats_sheet.max_row >= 2  # At least header + one data row

    def test_excel_export_korean_characters(self):
        """
        Test that Korean characters are properly exported.

        Expected: 홍길동 appears correctly in Excel
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        data_sheet = workbook.active

        # Find Korean text in data
        found_korean = False
        for row in data_sheet.iter_rows(min_row=2, max_row=data_sheet.max_row):
            for cell in row:
                if cell.value and '홍길동' in str(cell.value):
                    found_korean = True
                    break

        assert found_korean

    def test_excel_export_empty_dataset(self):
        """
        Test Excel export for dataset with no records.

        Expected: 200 OK with workbook containing headers only
        """
        empty_dataset = Dataset.objects.create(
            title='Empty Dataset',
            filename='empty.xlsx',
            file_size=512,
            record_count=0,
            uploaded_by=self.admin_user
        )

        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': empty_dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        assert workbook is not None

    def test_excel_export_nonexistent_dataset(self):
        """
        Test Excel export for non-existent dataset.

        Expected: 404 Not Found
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': 99999})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_excel_export_content_disposition(self):
        """
        @SPEC:REQ-EXPORT-003
        Test that Content-Disposition header is correct.

        Expected: Filename ends with .xlsx
        """
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        content_disposition = response['Content-Disposition']
        assert 'attachment' in content_disposition
        assert '.xlsx' in content_disposition

    def test_excel_export_null_values(self):
        """
        Test Excel export handling of null values.

        Expected: Null values appear as empty cells
        """
        DataRecord.objects.create(
            dataset=self.dataset,
            data={
                'student_name': 'Test User',
                'age': None,
                'gpa': 3.5,
                'enrolled': None,
                'category': 'enrollment'
            }
        )

        self.client.force_authenticate(user=self.admin_user)
        url = reverse('dataset-export-excel', kwargs={'pk': self.dataset.pk})
        response = self.client.post(url)

        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        data_sheet = workbook.active

        # Find row with Test User
        found_test_user = False
        for row in data_sheet.iter_rows(min_row=2):
            if row[0].value == 'Test User':
                found_test_user = True
                # Check that age cell is None or empty
                age_cell = row[1]  # Assuming age is second column
                assert age_cell.value is None or age_cell.value == ''
                break

        assert found_test_user
