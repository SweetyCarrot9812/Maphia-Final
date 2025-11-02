"""
Tests for Excel Parser Service

@SPEC:DASH-001
@TEST:EXCEL-PARSER

Tests the Excel file parsing functionality using openpyxl.
"""

import pytest
from io import BytesIO
from openpyxl import Workbook


@pytest.mark.django_db
@pytest.mark.unit
class TestExcelParserService:
    """
    Tests for ExcelParserService.

    Requirements:
    - REQ-DASH-001: Parse Excel files (.xlsx, .xls)
    - Validate file format and size
    - Extract data into structured records
    """

    def create_sample_excel(self, data_rows):
        """
        Helper: Create a sample Excel file in memory.

        Args:
            data_rows: List of lists representing Excel rows (first row is header)

        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        ws = wb.active

        for row in data_rows:
            ws.append(row)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    def test_parse_simple_excel_file(self):
        """
        @TEST:EXCEL-PARSER-001
        Parse a simple Excel file with headers and data rows
        """
        from dashboard.services.excel_parser import ExcelParserService

        data = [
            ['Name', 'Age', 'Department'],
            ['Alice', 25, 'Engineering'],
            ['Bob', 30, 'Marketing'],
            ['Charlie', 28, 'Sales']
        ]

        excel_file = self.create_sample_excel(data)

        parser = ExcelParserService()
        result = parser.parse(excel_file)

        assert result['success'] is True
        assert len(result['records']) == 3
        assert result['headers'] == ['Name', 'Age', 'Department']
        assert result['records'][0] == {'Name': 'Alice', 'Age': 25, 'Department': 'Engineering'}
        assert result['records'][1] == {'Name': 'Bob', 'Age': 30, 'Department': 'Marketing'}

    def test_parse_excel_with_empty_cells(self):
        """
        @TEST:EXCEL-PARSER-002
        Handle Excel files with empty cells
        """
        from dashboard.services.excel_parser import ExcelParserService

        data = [
            ['Name', 'Age', 'Department'],
            ['Alice', 25, 'Engineering'],
            ['Bob', None, 'Marketing'],  # Missing age
            ['Charlie', 28, None]  # Missing department
        ]

        excel_file = self.create_sample_excel(data)

        parser = ExcelParserService()
        result = parser.parse(excel_file)

        assert result['success'] is True
        assert len(result['records']) == 3
        assert result['records'][1] == {'Name': 'Bob', 'Age': None, 'Department': 'Marketing'}
        assert result['records'][2] == {'Name': 'Charlie', 'Age': 28, 'Department': None}

    def test_parse_excel_with_numeric_and_text_data(self):
        """
        @TEST:EXCEL-PARSER-003
        Parse mixed data types (numbers, text, dates)
        """
        from dashboard.services.excel_parser import ExcelParserService
        from datetime import datetime

        data = [
            ['Student ID', 'Name', 'GPA', 'Enrollment Date'],
            [1001, 'Alice', 3.8, datetime(2023, 9, 1)],
            [1002, 'Bob', 3.5, datetime(2023, 9, 1)],
        ]

        excel_file = self.create_sample_excel(data)

        parser = ExcelParserService()
        result = parser.parse(excel_file)

        assert result['success'] is True
        assert len(result['records']) == 2
        assert result['records'][0]['Student ID'] == 1001
        assert result['records'][0]['GPA'] == 3.8
        assert isinstance(result['records'][0]['Enrollment Date'], (str, datetime))

    def test_parse_excel_without_headers(self):
        """
        @TEST:EXCEL-PARSER-004
        Handle Excel files without explicit headers (use column letters)
        """
        from dashboard.services.excel_parser import ExcelParserService

        data = [
            ['Alice', 25, 'Engineering'],
            ['Bob', 30, 'Marketing'],
        ]

        excel_file = self.create_sample_excel(data)

        parser = ExcelParserService()
        result = parser.parse(excel_file, has_header=False)

        assert result['success'] is True
        assert len(result['records']) == 2
        # Should use column letters as keys
        assert 'A' in result['records'][0] or 'col_0' in result['records'][0]

    def test_parse_empty_excel_file(self):
        """
        @TEST:EXCEL-PARSER-005
        Handle empty Excel files gracefully
        """
        from dashboard.services.excel_parser import ExcelParserService

        data = []
        excel_file = self.create_sample_excel(data)

        parser = ExcelParserService()
        result = parser.parse(excel_file)

        assert result['success'] is False
        assert 'error' in result
        assert 'empty' in result['error'].lower()

    def test_parse_excel_with_only_headers(self):
        """
        @TEST:EXCEL-PARSER-006
        Handle Excel files with only header row (no data)
        """
        from dashboard.services.excel_parser import ExcelParserService

        data = [
            ['Name', 'Age', 'Department']
        ]

        excel_file = self.create_sample_excel(data)

        parser = ExcelParserService()
        result = parser.parse(excel_file)

        assert result['success'] is True
        assert len(result['records']) == 0
        assert result['headers'] == ['Name', 'Age', 'Department']

    def test_parse_invalid_file(self):
        """
        @TEST:EXCEL-PARSER-007
        Reject invalid files (not Excel format)
        """
        from dashboard.services.excel_parser import ExcelParserService

        invalid_file = BytesIO(b"This is not an Excel file")

        parser = ExcelParserService()
        result = parser.parse(invalid_file)

        assert result['success'] is False
        assert 'error' in result

    def test_parse_excel_with_multiple_sheets(self):
        """
        @TEST:EXCEL-PARSER-008
        Parse first sheet when multiple sheets exist
        """
        from dashboard.services.excel_parser import ExcelParserService

        wb = Workbook()

        # Sheet 1 (should be parsed)
        ws1 = wb.active
        ws1.title = "Students"
        ws1.append(['Name', 'Age'])
        ws1.append(['Alice', 25])

        # Sheet 2 (should be ignored)
        ws2 = wb.create_sheet("Grades")
        ws2.append(['Student', 'Grade'])
        ws2.append(['Alice', 'A'])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        parser = ExcelParserService()
        result = parser.parse(excel_file)

        assert result['success'] is True
        assert len(result['records']) == 1
        assert result['headers'] == ['Name', 'Age']
        assert result['records'][0] == {'Name': 'Alice', 'Age': 25}

    def test_get_file_metadata(self):
        """
        @TEST:EXCEL-PARSER-009
        Extract file metadata (sheet names, row count, column count)
        """
        from dashboard.services.excel_parser import ExcelParserService

        data = [
            ['Name', 'Age', 'Department'],
            ['Alice', 25, 'Engineering'],
            ['Bob', 30, 'Marketing'],
        ]

        excel_file = self.create_sample_excel(data)

        parser = ExcelParserService()
        metadata = parser.get_metadata(excel_file)

        assert metadata['sheet_count'] >= 1
        assert metadata['row_count'] == 3  # header + 2 data rows
        assert metadata['column_count'] == 3
