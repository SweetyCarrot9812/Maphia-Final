"""
Dashboard views for University Data Visualization Dashboard.

ViewSets:
- DatasetViewSet: CRUD operations for datasets with file upload
- DataRecordViewSet: CRUD operations for data records
- StatisticsViewSet: Analytics and aggregated data endpoints
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from django.db.models import Count, Sum
from django.shortcuts import get_object_or_404
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import csv
import re
from io import BytesIO, StringIO
from django.db.models import Count, Avg, Max, Min
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from .models import Dataset, DataRecord
from .serializers import (
    DatasetSerializer,
    DatasetListSerializer,
    DatasetCreateSerializer,
    DataRecordSerializer,
    DatasetStatisticsSerializer,
)
from .services.excel_parser import ExcelParserService


class DatasetViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Dataset model.

    Endpoints:
    - GET /datasets/ - List all datasets
    - POST /datasets/ - Create new dataset with file upload
    - GET /datasets/{id}/ - Retrieve dataset details
    - PUT /datasets/{id}/ - Update dataset
    - PATCH /datasets/{id}/ - Partial update dataset
    - DELETE /datasets/{id}/ - Delete dataset
    - POST /datasets/{id}/upload/ - Upload Excel file and parse
    - GET /datasets/{id}/records/ - Get all records for dataset
    """

    queryset = Dataset.objects.all().select_related('uploaded_by')
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return DatasetListSerializer
        elif self.action == 'create':
            return DatasetCreateSerializer
        return DatasetSerializer

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)

        return queryset

    def perform_create(self, serializer):
        """
        Create dataset and process uploaded file.

        Assigns current user as uploader and processes Excel file.
        """
        # Save dataset with current user
        dataset = serializer.save(uploaded_by=self.request.user)

        # Process uploaded file if present
        if 'file' in self.request.FILES:
            self._process_excel_file(dataset, self.request.FILES['file'])

    def _process_excel_file(self, dataset, uploaded_file):
        """
        Process uploaded Excel file and create DataRecords.

        Args:
            dataset: Dataset instance
            uploaded_file: Uploaded file object

        Returns:
            Number of records created
        """
        try:
            # Read Excel file using openpyxl
            file_content = uploaded_file.read()
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            sheet = workbook.active

            # Convert to pandas DataFrame for easier processing
            data = []
            headers = []

            # Get headers from first row
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else f"Column_{len(headers) + 1}")

            # Get data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        # Convert value to JSON-serializable format
                        if pd.isna(value):
                            row_dict[headers[idx]] = None
                        elif isinstance(value, (int, float, str, bool)):
                            row_dict[headers[idx]] = value
                        else:
                            row_dict[headers[idx]] = str(value)
                data.append(row_dict)

            # Create DataRecord instances in bulk
            records = [
                DataRecord(dataset=dataset, data=row_data)
                for row_data in data
            ]
            DataRecord.objects.bulk_create(records)

            # Update dataset record count
            dataset.record_count = len(records)
            dataset.save()

            return len(records)

        except Exception as e:
            # If processing fails, delete the dataset
            dataset.delete()
            raise Exception(f"Failed to process Excel file: {str(e)}")

    @action(detail=False, methods=['post'], url_path='upload')
    def upload_file(self, request):
        """
        Upload Excel file and create new dataset with parsed data.

        @CODE:DASH-API-UPLOAD

        Request:
            POST /api/datasets/upload/
            Body: multipart/form-data with:
                - file: Excel file
                - title: Dataset title
                - description: (optional)
                - category: (optional)

        Response:
            201: Dataset created with records
            400: Validation error or parsing error
            401: Unauthorized
        """
        # Validate request data
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if 'title' not in request.data:
            return Response(
                {"error": "Title is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        uploaded_file = request.FILES['file']

        # Validate file format
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {"error": "Invalid file format. Only .xlsx and .xls are supported."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate file size (10MB max)
        max_size = 10 * 1024 * 1024
        if uploaded_file.size > max_size:
            return Response(
                {"error": f"File size exceeds maximum of 10MB"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Parse Excel file using ExcelParserService
            parser = ExcelParserService()

            # Reset file pointer to beginning
            uploaded_file.seek(0)

            parse_result = parser.parse(uploaded_file)

            if not parse_result['success']:
                return Response(
                    {"error": parse_result['error']},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Create Dataset
            dataset = Dataset.objects.create(
                title=request.data['title'],
                description=request.data.get('description', ''),
                category=request.data.get('category', ''),
                filename=uploaded_file.name,
                file_size=uploaded_file.size,
                uploaded_by=request.user,
                record_count=len(parse_result['records'])
            )

            # Create DataRecords
            records_to_create = [
                DataRecord(dataset=dataset, data=record)
                for record in parse_result['records']
            ]
            DataRecord.objects.bulk_create(records_to_create)

            # Serialize response
            serializer = DatasetSerializer(dataset)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {"error": f"Failed to process file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def upload(self, request, pk=None):
        """
        Upload and process Excel file for existing dataset.

        Request:
            POST /datasets/{id}/upload/
            Body: multipart/form-data with 'file' field

        Response:
            {
                "status": "success",
                "records_created": 100,
                "message": "File processed successfully"
            }
        """
        dataset = self.get_object()

        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            records_created = self._process_excel_file(dataset, request.FILES['file'])

            return Response({
                "status": "success",
                "records_created": records_created,
                "message": "File processed successfully"
            })

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def records(self, request, pk=None):
        """
        Get all records for a dataset.

        Request:
            GET /datasets/{id}/records/?page=1&page_size=50

        Response:
            {
                "count": 100,
                "results": [...]
            }
        """
        dataset = self.get_object()
        records = dataset.records.all()

        # Apply pagination
        page = self.paginate_queryset(records)
        if page is not None:
            serializer = DataRecordSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = DataRecordSerializer(records, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='export/csv', permission_classes=[IsAuthenticated])
    def export_csv(self, request, pk=None):
        """
        Export dataset to CSV format with UTF-8 BOM encoding.

        @SPEC:REQ-EXPORT-001 - CSV Export Basic Functionality
        @SPEC:REQ-EXPORT-002 - RFC 4180 Special Character Handling

        Request:
            POST /api/datasets/{id}/export/csv/

        Response:
            200: CSV file download
            403: Permission denied (Viewer role)
            404: Dataset not found

        Features:
        - UTF-8 encoding with BOM for Excel compatibility
        - RFC 4180 compliant (proper escaping of commas, quotes, newlines)
        - Content-Disposition header for browser download
        - Role-based access: Admin/Manager only
        """
        # Check user role (Admin or Manager only)
        if hasattr(request.user, 'role') and request.user.role not in ['admin', 'manager']:
            return Response(
                {"error": "Only Admin and Manager users can export data."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get dataset
        dataset = self.get_object()

        # Get all records for this dataset
        records = dataset.records.all()

        # Prepare CSV data
        if records.exists():
            # Extract field names from first record
            first_record = records.first()
            fieldnames = list(first_record.data.keys())

            # Create CSV in memory with UTF-8-sig (includes BOM)
            output = StringIO()
            writer = csv.DictWriter(
                output,
                fieldnames=fieldnames,
                quoting=csv.QUOTE_MINIMAL,  # RFC 4180 compliance
                lineterminator='\n'
            )

            # Write header
            writer.writeheader()

            # Write data rows
            for record in records:
                # Convert None to empty string for CSV
                row_data = {
                    k: (v if v is not None else '')
                    for k, v in record.data.items()
                }
                writer.writerow(row_data)

            csv_content = output.getvalue()
        else:
            # Empty dataset - just headers
            csv_content = ""

        # Sanitize filename
        safe_filename = self._sanitize_filename(dataset.title)
        filename = f"{safe_filename}.csv"

        # Create HTTP response with proper headers
        response = HttpResponse(
            content_type='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )

        # Write BOM for UTF-8
        response.write('\ufeff')  # UTF-8 BOM

        # Write CSV content
        response.write(csv_content)

        return response

    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitize filename by removing special characters.

        Args:
            filename: Original filename

        Returns:
            Sanitized filename safe for use in Content-Disposition header
        """
        # Replace spaces with underscores
        filename = filename.replace(' ', '_')

        # Remove special characters, keep alphanumeric, underscore, hyphen
        filename = re.sub(r'[^a-zA-Z0-9_\-가-힣]', '', filename)

        # Limit length
        if len(filename) > 200:
            filename = filename[:200]

        return filename or 'dataset'

    @action(detail=True, methods=['post'], url_path='export/excel', permission_classes=[IsAuthenticated])
    def export_excel(self, request, pk=None):
        """
        Export dataset to Excel format with multiple sheets and styling.

        @SPEC:REQ-EXPORT-003 - Excel Basic Export with Styling
        @SPEC:REQ-EXPORT-004 - Excel Multi-sheet Support

        Request:
            POST /api/datasets/{id}/export/excel/

        Response:
            200: Excel file download
            403: Permission denied (Viewer role)
            404: Dataset not found

        Features:
        - Multi-sheet workbook (Data, Statistics, Charts)
        - Header styling (bold, blue background, white text)
        - Auto-width columns
        - Zebra striping (alternating row colors)
        - Role-based access: Admin/Manager only
        """
        # Check user role (Admin or Manager only)
        if hasattr(request.user, 'role') and request.user.role not in ['admin', 'manager']:
            return Response(
                {"error": "Only Admin and Manager users can export data."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get dataset
        dataset = self.get_object()

        # Create workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Sheet 1: 데이터 (Data)
        self._create_data_sheet(workbook, dataset)

        # Sheet 2: 요약 통계 (Summary Statistics)
        self._create_statistics_sheet(workbook, dataset)

        # Save to BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Sanitize filename
        safe_filename = self._sanitize_filename(dataset.title)
        filename = f"{safe_filename}.xlsx"

        # Create HTTP response
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )

        return response

    def _create_data_sheet(self, workbook, dataset):
        """
        Create data sheet with all records.

        @SPEC:REQ-EXPORT-003

        Features:
        - Header row with styling (bold, blue background, white text)
        - Auto-width columns
        - Zebra striping (alternating row colors)
        """
        # Create sheet
        data_sheet = workbook.create_sheet(title='데이터')

        # Get records
        records = dataset.records.all()

        if not records.exists():
            # Empty dataset - just add a note
            data_sheet['A1'] = '데이터 없음'
            return

        # Get field names from first record
        first_record = records.first()
        fieldnames = list(first_record.data.keys())

        # Define header styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Write header row
        for col_idx, field_name in enumerate(fieldnames, start=1):
            cell = data_sheet.cell(row=1, column=col_idx)
            cell.value = field_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Define zebra striping styles
        even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            for col_idx, field_name in enumerate(fieldnames, start=1):
                cell = data_sheet.cell(row=row_idx, column=col_idx)
                value = record.data.get(field_name)

                # Convert value to appropriate type
                if value is None:
                    cell.value = ''
                elif isinstance(value, bool):
                    cell.value = str(value)
                else:
                    cell.value = value

                # Apply zebra striping to even rows
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        # Auto-adjust column widths
        for col_idx, field_name in enumerate(fieldnames, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(field_name))

            # Check data cell lengths
            for row_idx in range(2, data_sheet.max_row + 1):
                cell_value = data_sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set column width (add padding)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            data_sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_statistics_sheet(self, workbook, dataset):
        """
        Create statistics sheet with aggregated data.

        @SPEC:REQ-EXPORT-004

        Includes:
        - Total record count
        - Category breakdown
        - Numeric field statistics (avg, min, max)
        """
        # Create sheet
        stats_sheet = workbook.create_sheet(title='요약 통계')

        # Define header styles
        header_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=14, color='4472C4')

        # Title
        stats_sheet['A1'] = f'{dataset.title} - 요약 통계'
        stats_sheet['A1'].font = title_font
        stats_sheet.merge_cells('A1:B1')

        # Dataset info
        row = 3
        stats_sheet[f'A{row}'] = '데이터셋 정보'
        stats_sheet[f'A{row}'].font = header_font
        row += 1

        stats_sheet[f'A{row}'] = '총 레코드 수'
        stats_sheet[f'B{row}'] = dataset.record_count
        row += 1

        stats_sheet[f'A{row}'] = '파일 크기'
        stats_sheet[f'B{row}'] = f'{dataset.file_size / 1024:.2f} KB'
        row += 1

        stats_sheet[f'A{row}'] = '업로드 날짜'
        stats_sheet[f'B{row}'] = dataset.upload_date.strftime('%Y-%m-%d %H:%M')
        row += 2

        # Numeric field statistics
        records = dataset.records.all()
        if records.exists():
            stats_sheet[f'A{row}'] = '필드 통계'
            stats_sheet[f'A{row}'].font = header_font
            row += 1

            # Get numeric fields from first record
            first_record = records.first()
            numeric_fields = []

            for field_name, value in first_record.data.items():
                if isinstance(value, (int, float)) and value is not None:
                    numeric_fields.append(field_name)

            # Calculate statistics for numeric fields
            if numeric_fields:
                # Header row for statistics
                stats_sheet[f'A{row}'] = '필드'
                stats_sheet[f'B{row}'] = '평균'
                stats_sheet[f'C{row}'] = '최소'
                stats_sheet[f'D{row}'] = '최대'
                for col in ['A', 'B', 'C', 'D']:
                    stats_sheet[f'{col}{row}'].font = Font(bold=True)
                row += 1

                for field_name in numeric_fields:
                    values = [
                        record.data.get(field_name)
                        for record in records
                        if record.data.get(field_name) is not None
                    ]

                    if values:
                        avg_val = sum(values) / len(values)
                        min_val = min(values)
                        max_val = max(values)

                        stats_sheet[f'A{row}'] = field_name
                        stats_sheet[f'B{row}'] = round(avg_val, 2)
                        stats_sheet[f'C{row}'] = min_val
                        stats_sheet[f'D{row}'] = max_val
                        row += 1

        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            stats_sheet.column_dimensions[col].width = 20

    @action(detail=True, methods=['post'], url_path='export/pdf', permission_classes=[IsAuthenticated])
    def export_pdf(self, request, pk=None):
        """
        Export dataset to PDF format with tables and formatting.

        @SPEC:REQ-EXPORT-005 - PDF Basic Export
        @SPEC:REQ-EXPORT-006 - PDF Styling and Layout

        Request:
            POST /api/datasets/{id}/export/pdf/

        Response:
            200: PDF file download
            403: Permission denied (Viewer role)
            404: Dataset not found

        Features:
        - Table formatting with headers
        - Summary statistics section
        - Header/footer with branding
        - Pagination
        - Timestamp
        - Role-based access: Admin/Manager only
        """
        # Check user role (Admin or Manager only)
        if hasattr(request.user, 'role') and request.user.role not in ['admin', 'manager']:
            return Response(
                {"error": "Only Admin and Manager users can export data."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get dataset
        dataset = self.get_object()

        # Create PDF in memory
        pdf_buffer = BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=50
        )

        # Container for PDF elements
        elements = []

        # Get styles
        styles = getSampleStyleSheet()

        # Custom title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # Custom heading style
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Add title
        title = Paragraph(f"<b>{dataset.title}</b>", title_style)
        elements.append(title)

        # Add export timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        timestamp_text = Paragraph(
            f"<i>Generated on: {timestamp}</i>",
            styles['Normal']
        )
        elements.append(timestamp_text)
        elements.append(Spacer(1, 20))

        # Add summary statistics section
        elements.append(Paragraph("<b>Dataset Summary</b>", heading_style))

        summary_data = [
            ['Description:', dataset.description or 'N/A'],
            ['Category:', dataset.category or 'N/A'],
            ['Total Records:', str(dataset.record_count)],
            ['File Size:', f'{dataset.file_size / 1024:.2f} KB'],
            ['Upload Date:', dataset.upload_date.strftime('%Y-%m-%d %H:%M')],
            ['Uploaded By:', dataset.uploaded_by.username],
        ]

        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Add data table section
        records = dataset.records.all()

        if records.exists():
            elements.append(Paragraph("<b>Data Records</b>", heading_style))

            # Get field names from first record
            first_record = records.first()
            fieldnames = list(first_record.data.keys())

            # Prepare table data
            table_data = [fieldnames]  # Header row

            # Add data rows
            for record in records:
                row = []
                for field_name in fieldnames:
                    value = record.data.get(field_name)
                    if value is None:
                        row.append('')
                    else:
                        row.append(str(value))
                table_data.append(row)

            # Calculate column widths dynamically
            available_width = 6.5 * inch
            num_columns = len(fieldnames)
            col_width = available_width / num_columns

            # Create table
            data_table = Table(table_data, colWidths=[col_width] * num_columns, repeatRows=1)

            # Apply styling
            data_table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

                # Data rows styling
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

                # Zebra striping
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))

            elements.append(data_table)
        else:
            # No records message
            no_data_text = Paragraph(
                "<i>No data records available for this dataset.</i>",
                styles['Normal']
            )
            elements.append(no_data_text)

        # Build PDF
        doc.build(elements)

        # Get PDF content
        pdf_buffer.seek(0)
        pdf_content = pdf_buffer.getvalue()
        pdf_buffer.close()

        # Sanitize filename
        safe_filename = self._sanitize_filename(dataset.title)
        filename = f"{safe_filename}.pdf"

        # Create HTTP response
        response = HttpResponse(
            pdf_content,
            content_type='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )

        return response


class DataRecordViewSet(viewsets.ModelViewSet):
    """
    ViewSet for DataRecord model.

    Endpoints:
    - GET /records/ - List all records
    - POST /records/ - Create new record
    - GET /records/{id}/ - Retrieve record
    - PUT /records/{id}/ - Update record
    - PATCH /records/{id}/ - Partial update record
    - DELETE /records/{id}/ - Delete record
    """

    queryset = DataRecord.objects.all().select_related('dataset')
    serializer_class = DataRecordSerializer

    def get_queryset(self):
        """Filter records by dataset if dataset_id is provided."""
        queryset = super().get_queryset()

        dataset_id = self.request.query_params.get('dataset_id')
        if dataset_id:
            queryset = queryset.filter(dataset_id=dataset_id)

        return queryset


class StatisticsViewSet(viewsets.ViewSet):
    """
    ViewSet for dashboard statistics and analytics.

    Endpoints:
    - GET /statistics/overview/ - Dashboard overview statistics
    """

    @action(detail=False, methods=['get'])
    def overview(self, request):
        """
        Get dashboard overview statistics.

        Response:
            {
                "total_datasets": 10,
                "total_records": 1000,
                "total_size": 5242880,
                "categories": [
                    {"category": "enrollment", "count": 5},
                    {"category": "grades", "count": 3}
                ],
                "recent_uploads": [...]
            }
        """
        # Aggregate statistics
        stats = Dataset.objects.aggregate(
            total_datasets=Count('id'),
            total_records=Sum('record_count'),
            total_size=Sum('file_size')
        )

        # Category breakdown
        categories = (
            Dataset.objects
            .values('category')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        # Recent uploads (last 5)
        recent_uploads = Dataset.objects.all()[:5]

        # Serialize response
        data = {
            'total_datasets': stats['total_datasets'] or 0,
            'total_records': stats['total_records'] or 0,
            'total_size': stats['total_size'] or 0,
            'categories': list(categories),
            'recent_uploads': DatasetListSerializer(recent_uploads, many=True).data
        }

        serializer = DatasetStatisticsSerializer(data=data)
        serializer.is_valid(raise_exception=True)

        return Response(serializer.data)
